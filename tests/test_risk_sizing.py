"""Tests for the risk-contribution sizing tool.

The failure that matters here is silent and directional: a weighting bug that
sizes *up* into volatility rather than down would look like a plausible table
and quietly concentrate risk. So the sign of the relationship is asserted
explicitly, not just the arithmetic.
"""

from __future__ import annotations

import datetime as dt
import math
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))
import risk_sizing as rs  # noqa: E402


def price_series(start: float, weekly: list[float]) -> dict:
    """{date: price} walking `start` forward by the given weekly returns."""
    d, p, out = dt.datetime(2025, 1, 6), start, {}
    out[d] = p
    for r in weekly:
        d += dt.timedelta(days=7)
        p *= (1 + r)
        out[d] = p
    return out


def alternating(magnitude: float, n: int = 40) -> list[float]:
    """Zero-drift saw-tooth: realized vol scales with `magnitude`."""
    return [magnitude if i % 2 == 0 else -magnitude for i in range(n)]


def build_workbook(path: Path, holdings, vols: dict[str, float],
                   account: str = "Joint Brokerage") -> Path:
    """`holdings` is {ticker: value}, or {account: {ticker: value}}."""
    if holdings and not isinstance(next(iter(holdings.values())), dict):
        holdings = {account: holdings}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset PM"

    # ACCOUNT SUMMARY block — the parser learns real account names here, so
    # asset-class labels sharing column A are not mistaken for accounts.
    hdr = [None] * 12
    hdr[rs.ACCOUNT_NAME_COL] = "Account"
    ws.append(hdr)
    for acct in holdings:
        row = [None] * 12
        row[rs.ACCOUNT_NAME_COL] = acct
        ws.append(row)
    tot = [None] * 12
    tot[rs.ACCOUNT_NAME_COL] = "TOTAL PORTFOLIO"
    ws.append(tot)

    for acct, positions in holdings.items():
        ws.append([acct] + [None] * 11)          # account section header
        ws.append(["US Stocks"] + [None] * 11)   # asset-class label, not an account
        for ticker, value in positions.items():
            row = [None] * 12
            row[rs.HOLD_TICKER_COL] = ticker
            row[rs.HOLD_VALUE_COL] = value
            ws.append(row)
    for ticker, mag in vols.items():
        sh = wb.create_sheet(f"HMM {ticker}")
        sh.append([None] * 10)
        for d, p in price_series(100.0, alternating(mag)).items():
            row = [None] * 10
            row[rs.DATE_COL] = d
            row[rs.PRICE_COL] = p
            sh.append(row)
    wb.save(path)
    return path


@pytest.fixture
def book(tmp_path):
    return build_workbook(
        tmp_path / "p.xlsx",
        holdings={"CALM": 50_000, "WILD": 50_000, "CASHX": 10_000},
        vols={"CALM": 0.01, "WILD": 0.04},
    )


def load(book):
    wb = openpyxl.load_workbook(book, data_only=True, read_only=True)
    return rs.build_report(rs.load_holdings(wb), rs.load_price_history(wb, min_weeks=20))


# ------------------------------------------------------------ parsing

def test_pooled_holdings_sum_the_same_name_across_accounts(tmp_path):
    p = build_workbook(tmp_path / "d.xlsx",
                       {"401K": {"AAA": 1_000}, "Joint Brokerage": {"AAA": 2_500}},
                       {"AAA": 0.02})
    assert rs.load_holdings(openpyxl.load_workbook(p, data_only=True))["AAA"] == 3_500


def test_accounts_are_kept_separate(tmp_path):
    p = build_workbook(tmp_path / "a.xlsx",
                       {"401K": {"AAA": 1_000}, "Joint Brokerage": {"BBB": 2_500}},
                       {"AAA": 0.02, "BBB": 0.03})
    by = rs.load_holdings_by_account(openpyxl.load_workbook(p, data_only=True))
    assert by == {"401K": {"AAA": 1_000.0}, "Joint Brokerage": {"BBB": 2_500.0}}


def test_asset_class_labels_are_not_treated_as_accounts(tmp_path):
    """"US Stocks" sits alone in column A exactly like an account header."""
    p = build_workbook(tmp_path / "l.xlsx", {"401K": {"AAA": 100}}, {"AAA": 0.02})
    by = rs.load_holdings_by_account(openpyxl.load_workbook(p, data_only=True))
    assert set(by) == {"401K"}, f"asset-class label leaked in: {set(by)}"


@pytest.mark.parametrize("name,expected", [
    ("Brokerage Link (401K)", True), ("ROTH IRA", True),
    ("Health Savings Account (HSA)", True), ("Joint Brokerage", False),
    ("Tre Brokerage", False), ("Crypto", False),
])
def test_tax_status_inferred_from_account_name(name, expected):
    assert rs.is_deferred(name, set(), set()) is expected


def test_tax_status_can_be_overridden(tmp_path):
    assert rs.is_deferred("Joint Brokerage", set(), {"Joint Brokerage"}) is True
    assert rs.is_deferred("ROTH IRA", {"ROTH IRA"}, set()) is False


def test_holdings_without_price_history_are_reported_not_sized(book):
    rep = load(book)
    assert "CASHX" in rep["no_history"]
    assert all(r["ticker"] != "CASHX" for r in rep["rows"])


def test_summary_sheets_are_not_mistaken_for_price_sheets(tmp_path):
    p = build_workbook(tmp_path / "s.xlsx", {"AAA": 1}, {"AAA": 0.02})
    wb = openpyxl.load_workbook(p)
    wb.create_sheet("HMM Data")
    wb.create_sheet("HMM Backtest")
    wb.save(p)
    assert set(rs.load_price_history(openpyxl.load_workbook(p, data_only=True),
                                     min_weeks=20)) == {"AAA"}


# ------------------------------------------------------------ the math

def test_volatility_is_annualized(tmp_path):
    p = build_workbook(tmp_path / "v.xlsx", {"AAA": 1}, {"AAA": 0.02})
    hist = rs.load_price_history(openpyxl.load_workbook(p, data_only=True), min_weeks=20)
    vol = rs.annualized_vol(hist["AAA"], window=13)
    assert vol == pytest.approx(0.02 * math.sqrt(52), rel=0.10)


def test_higher_volatility_gets_a_smaller_target_weight(book):
    rep = load(book)
    by = {r["ticker"]: r for r in rep["rows"]}
    assert by["WILD"]["vol"] > by["CALM"]["vol"]
    assert by["WILD"]["target_weight"] < by["CALM"]["target_weight"], (
        "sizing up into volatility — the sign is inverted"
    )


def test_target_weights_are_inverse_proportional_to_volatility(book):
    rep = load(book)
    by = {r["ticker"]: r for r in rep["rows"]}
    ratio = by["CALM"]["target_weight"] / by["WILD"]["target_weight"]
    assert ratio == pytest.approx(by["WILD"]["vol"] / by["CALM"]["vol"], rel=0.05)


def test_weights_and_risk_shares_each_sum_to_one(book):
    rep = load(book)
    assert sum(r["target_weight"] for r in rep["rows"]) == pytest.approx(1.0)
    assert sum(r["weight"] for r in rep["rows"]) == pytest.approx(1.0)
    assert sum(r["risk_share"] for r in rep["rows"]) == pytest.approx(1.0)


def test_equal_capital_in_unequal_vol_gives_unequal_risk(book):
    """The mismatch the tool exists to expose."""
    rep = load(book)
    by = {r["ticker"]: r for r in rep["rows"]}
    assert by["CALM"]["weight"] == pytest.approx(by["WILD"]["weight"])
    assert by["WILD"]["risk_share"] > 3 * by["CALM"]["risk_share"]


def test_trades_net_to_zero_and_move_toward_target(book):
    rep = load(book)
    assert sum(r["trade"] for r in rep["rows"]) == pytest.approx(0.0, abs=1e-6)
    for r in rep["rows"]:
        if abs(r["target_weight"] - r["weight"]) > 1e-9:
            assert (r["trade"] > 0) == (r["target_weight"] > r["weight"])


def test_rows_are_ordered_by_risk_share(book):
    shares = [r["risk_share"] for r in load(book)["rows"]]
    assert shares == sorted(shares, reverse=True)


def test_already_risk_balanced_book_needs_no_trades(tmp_path):
    """Capital already inverse to vol -> trades vanish."""
    p = build_workbook(tmp_path / "b.xlsx",
                       holdings={"CALM": 80_000, "WILD": 20_000},
                       vols={"CALM": 0.01, "WILD": 0.04})
    rep = load(p)
    for r in rep["rows"]:
        assert abs(r["trade"]) < 0.02 * rep["total"]


# ------------------------------------------------------------ output

def test_render_reports_the_top_concentration(book):
    text = rs.render(load(book))
    assert "WILD" in text
    assert "Largest: WILD" in text
    assert "of risk" in text


def test_render_labels_tax_status(book):
    rep = load(book)
    assert "tax-deferred" in rs.render({**rep, "deferred": True})
    assert "taxable" in rs.render({**rep, "deferred": False})


def test_no_forecast_caveat_is_in_the_footer():
    assert "no directional forecast" in rs.FOOTER


def test_report_is_json_serializable(book):
    import json
    json.loads(json.dumps(load(book), default=str))
