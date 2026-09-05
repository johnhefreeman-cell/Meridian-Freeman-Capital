"""Trend state against the 200-day simple moving average.

A name trading at or above its 200-day SMA is tagged **STRONG**; below it,
**WEAK**. That is the whole rule. This module computes it, and reports the
three things that decide whether the tag means anything on a given name:
distance from the average, whether the average itself is rising or falling,
and how long the name has held its current side.

**The weekly approximation, stated up front.** The workbook carries weekly
closes, not daily. 200 trading days is 40 weeks, and a 40-week SMA of weekly
closes is the standard equivalent — but it is an approximation, not the same
number. The two track within roughly a percent on a liquid name, which is
immaterial for a name 20% clear of its average and decisive for one sitting on
it. Any name inside `NEAR_BAND` of the line is flagged, because a true daily
SMA could put it on the other side.

**What the tag is good for, tested rather than assumed.** Walk-forward over
220 weeks and 20 names, the STRONG bucket did not out-return the WEAK bucket at
any horizon from one week to one year; the spread's sign flips with the window
length, which is what no signal looks like. What the tag does separate, and
separates consistently, is volatility: names below their 200-day average went
on to realize about 1.2x the volatility of names above it. Treat it as a risk
tag that informs position size, not as a selection rule that picks names. See
`docs/trend-strength.md` for the full test.

Usage:
    uv run python scripts/trend_state.py --workbook path/to/Portfolio.xlsx
    uv run python scripts/trend_state.py --workbook p.xlsx --json out.json
"""

from __future__ import annotations

import argparse
import json
import math
import statistics as st
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

# 200 trading days / 5 sessions per week. Fixed by the definition of the rule,
# not fitted to anything.
SMA_WEEKS = 40
TRADING_DAYS = 200
SLOPE_LOOKBACK = 13          # one quarter, for the direction of the average
NEAR_BAND = 0.03             # inside this, the daily/weekly gap could flip the tag
WEEKS_PER_YEAR = 52.0

PRICE_PREFIX = "HMM "
PRICE_SKIP = {"HMM Data", "HMM Backtest"}
DATE_COL, PRICE_COL = 7, 8

STRONG, WEAK = "STRONG", "WEAK"


def load_series(wb, min_weeks: int = SMA_WEEKS + 4) -> dict[str, dict]:
    """{ticker: {date: close}} from the per-name weekly price tabs."""
    out: dict[str, dict] = {}
    for sheet in wb.sheetnames:
        if not sheet.startswith(PRICE_PREFIX) or sheet in PRICE_SKIP:
            continue
        series = {}
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if len(row) <= PRICE_COL:
                continue
            date, price = row[DATE_COL], row[PRICE_COL]
            if date is not None and isinstance(price, (int, float)) and price > 0:
                series[date] = float(price)
        if len(series) >= min_weeks:
            out[sheet[len(PRICE_PREFIX):]] = {d: series[d] for d in sorted(series)}
    return out


def sma(closes: list[float], end: int, window: int = SMA_WEEKS) -> float | None:
    """Mean of `window` closes ending at index `end`, or None if short."""
    if end + 1 < window:
        return None
    return sum(closes[end + 1 - window:end + 1]) / window


def weeks_on_side(closes: list[float], strong: bool,
                  window: int = SMA_WEEKS) -> int:
    """Consecutive weeks the name has closed on its current side of the average.

    A name one week past a crossover and a name three years above the line get
    the same tag, so the tag alone hides the distinction that matters most.
    """
    run = 0
    for i in range(len(closes) - 1, window - 2, -1):
        m = sma(closes, i, window)
        if m is None:
            break
        if (closes[i] >= m) is strong:
            run += 1
        else:
            break
    return run


def classify(series: dict, window: int = SMA_WEEKS,
             price_override: float | None = None) -> dict | None:
    """Trend state for one name.

    `price_override` marks the name against a fresher price than the series
    carries — a brokerage quote from a later date, say — while the average
    stays as of the last weekly close. The result is labelled `stale_sma` so
    the mismatch is never silent.
    """
    dates = sorted(series)
    closes = [series[d] for d in dates]
    n = len(closes)
    avg = sma(closes, n - 1, window)
    if avg is None:
        return None

    price = price_override if price_override is not None else closes[-1]
    strong = price >= avg
    prev = sma(closes, n - 1 - SLOPE_LOOKBACK, window)

    return dict(
        state=STRONG if strong else WEAK,
        price=price,
        sma=avg,
        distance=price / avg - 1,
        sma_slope=(avg / prev - 1) if prev else None,
        sma_rising=(avg > prev) if prev else None,
        weeks_in_state=weeks_on_side(closes, closes[-1] >= avg, window),
        near_line=abs(price / avg - 1) < NEAR_BAND,
        as_of=dates[-1],
        weeks_of_history=n,
        stale_sma=price_override is not None,
    )


def classify_all(series_by_ticker: dict[str, dict],
                 window: int = SMA_WEEKS,
                 prices: dict[str, float] | None = None) -> dict[str, dict]:
    out = {}
    for ticker, series in series_by_ticker.items():
        got = classify(series, window,
                       (prices or {}).get(ticker))
        if got:
            out[ticker] = got
    return out


def forward_vol(series: dict, weeks: int = SLOPE_LOOKBACK) -> float | None:
    """Annualized volatility of the last `weeks` weekly returns."""
    dates = sorted(series)[-(weeks + 1):]
    if len(dates) < weeks + 1:
        return None
    rets = [series[dates[i]] / series[dates[i - 1]] - 1 for i in range(1, len(dates))]
    return st.pstdev(rets) * math.sqrt(WEEKS_PER_YEAR)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--window", type=int, default=SMA_WEEKS,
                    help=f"SMA length in weeks (default {SMA_WEEKS}, "
                         f"the {TRADING_DAYS}-day equivalent)")
    ap.add_argument("--json")
    ap.add_argument("--current-prices", action="store_true",
                    help="mark names against the holdings sheet's live prices "
                         "instead of the last weekly close; the average stays "
                         "as of the last weekly close and the result says so")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=True, read_only=True)
    series = load_series(wb)

    prices = None
    if args.current_prices:
        import portfolio_extract as px
        rows = list(wb[px.HOLDINGS_SHEET].iter_rows(values_only=True))
        accounts, _ = px.account_summary(rows)
        prices = {}
        for h in px.load_holdings(rows, {a["account"] for a in accounts}):
            if h.get("price"):
                prices.setdefault(h["ticker"], h["price"])

    states = classify_all(series, args.window, prices)
    if not states:
        raise SystemExit("no name has enough price history for the window")

    for t, s in states.items():
        s["vol"] = forward_vol(series[t])

    if args.json:
        with open(args.json, "w") as fh:
            json.dump(states, fh, indent=1, default=str)

    ordered = sorted(states.items(), key=lambda kv: -kv[1]["distance"])
    as_of = ordered[0][1]["as_of"]
    stale = any(s["stale_sma"] for s in states.values())
    print(f"Trend state vs the {args.window}-week "
          f"(~{args.window * 5}-day) SMA, as of {as_of:%Y-%m-%d}")
    if stale:
        print("Prices are the holdings sheet's; the averages are as of the last "
              "weekly close above.")
    print()
    print(f"{'name':<8}{'state':>7}{'price':>10}{'SMA':>10}{'vs SMA':>9}"
          f"{'SMA 13wk':>10}{'weeks':>7}{'13wk vol':>10}")
    print("-" * 71)
    for t, s in ordered:
        print(f"{t:<8}{s['state']:>7}{s['price']:>10.2f}{s['sma']:>10.2f}"
              f"{s['distance']:>+9.1%}"
              f"{(s['sma_slope'] or 0):>+10.1%}{s['weeks_in_state']:>7}"
              f"{(s['vol'] or 0):>10.0%}")

    strong = [t for t, s in states.items() if s["state"] == STRONG]
    near = [(t, s) for t, s in ordered if s["near_line"]]
    print(f"\n{len(strong)} strong / {len(states) - len(strong)} weak "
          f"of {len(states)} names with enough history.")
    sv = [s["vol"] for s in states.values() if s["state"] == STRONG and s["vol"]]
    wv = [s["vol"] for s in states.values() if s["state"] == WEAK and s["vol"]]
    if sv and wv:
        print(f"Median 13-week volatility — strong {st.median(sv):.0%}, "
              f"weak {st.median(wv):.0%}.")
    if near:
        print(f"\nWithin {NEAR_BAND:.0%} of the line, so a true daily SMA could "
              f"flip the tag:\n  " +
              ", ".join(f"{t} {s['distance']:+.1%}" for t, s in near))


if __name__ == "__main__":
    main()
