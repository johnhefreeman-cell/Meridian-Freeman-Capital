"""Risk-contribution and inverse-volatility sizing for the single-name sleeve.

Capital weight is not risk weight. A position at 27% of an account and 121%
annualized volatility carries most of that account's risk, and nothing in a
normal holdings view shows it. This reports both, and the trades that would
equalize them.

**Size per account, not across the portfolio.** Accounts cannot transfer to
each other, so a target computed over the whole sleeve is not executable — it
will tell you to sell in a 401(k) and buy in a taxable account with the
proceeds. `--by-account` is the default for that reason. It also separates the
trades that are free from the ones that cost tax, which usually matters more
than the sizing itself.

**This sizes positions. It does not time them.** Tested walk-forward on five
years of weekly data, direction-timing rules (a 2-state HMM, 40-week trend,
12-1 momentum) all captured well under buy-and-hold. Inverse-vol *sizing*
moved portfolio Sharpe 0.93 -> 0.97 and max drawdown -40% -> -35%: modest,
inside the noise for a correlated book, and it gave up wealth versus equal
weight. Treat it as risk hygiene, not alpha.

Workbook layout expected (matches IRA_Portfolio.v2.0.xlsx):
  - one sheet per name titled "HMM <TICKER>", weekly dates in column H and
    prices in column I
  - a holdings sheet ("Asset PM") whose ACCOUNT SUMMARY block lists account
    names in column B; each account section then repeats that name alone in
    column A, with tickers in column E and market value in column K

Usage:
    uv run python scripts/risk_sizing.py --workbook path/to/Portfolio.xlsx
    uv run python scripts/risk_sizing.py --workbook p.xlsx --pooled
    uv run python scripts/risk_sizing.py --workbook p.xlsx --json
"""

from __future__ import annotations

import argparse
import json
import math
import re
import statistics as st
import sys
from typing import Any

WEEKS_PER_YEAR = 52.0
TICKER_RE = re.compile(r"[A-Z]{1,5}(\.[A-Z])?")

PRICE_SHEET_PREFIX = "HMM "
PRICE_SHEET_SKIP = {"HMM Data", "HMM Backtest"}
DATE_COL, PRICE_COL = 7, 8
HOLD_TICKER_COL, HOLD_VALUE_COL = 4, 10
ACCOUNT_NAME_COL = 1

# Account names containing these are treated as tax-advantaged, so rebalancing
# inside them realizes nothing. Inferred from the name — override with
# --taxable / --deferred if your naming differs.
DEFERRED_HINTS = ("401K", "IRA", "HSA", "ROTH", "403B", "457", "SEP")


def load_price_history(wb, min_weeks: int = 60) -> dict[str, dict]:
    """{ticker: {date: price}} from every per-name price sheet."""
    out: dict[str, dict] = {}
    for name in wb.sheetnames:
        if not name.startswith(PRICE_SHEET_PREFIX) or name in PRICE_SHEET_SKIP:
            continue
        series = {}
        for row in wb[name].iter_rows(min_row=2, values_only=True):
            if len(row) <= PRICE_COL:
                continue
            date, price = row[DATE_COL], row[PRICE_COL]
            if date is not None and isinstance(price, (int, float)) and price > 0:
                series[date] = float(price)
        if len(series) >= min_weeks:
            out[name[len(PRICE_SHEET_PREFIX):]] = series
    return out


def account_names(rows: list[tuple]) -> list[str]:
    """Canonical account names from the ACCOUNT SUMMARY block.

    Needed because asset-class labels ("US Stocks", "Commodities") also sit
    alone in column A and are otherwise indistinguishable from an account
    header.
    """
    names, started = [], False
    for row in rows[:40]:
        if len(row) <= ACCOUNT_NAME_COL:
            continue
        cell = row[ACCOUNT_NAME_COL]
        if cell == "Account":
            started = True
            continue
        if started and isinstance(cell, str) and cell.strip():
            if cell.strip().upper().startswith("TOTAL"):
                break
            names.append(cell.strip())
    return names


def load_holdings_by_account(wb, sheet: str = "Asset PM") -> dict[str, dict[str, float]]:
    """{account: {ticker: market value}}."""
    if sheet not in wb.sheetnames:
        raise SystemExit(f"holdings sheet {sheet!r} not in workbook")
    rows = list(wb[sheet].iter_rows(values_only=True))
    known = set(account_names(rows))
    out: dict[str, dict[str, float]] = {}
    current = None
    for row in rows:
        if row and isinstance(row[0], str) and row[0].strip() in known \
                and all(c is None for c in row[1:]):
            current = row[0].strip()
            continue
        if len(row) <= HOLD_VALUE_COL:
            continue
        ticker, value = row[HOLD_TICKER_COL], row[HOLD_VALUE_COL]
        if not isinstance(ticker, str) or not isinstance(value, (int, float)):
            continue
        ticker = ticker.strip()
        if TICKER_RE.fullmatch(ticker) and value > 0 and current:
            bucket = out.setdefault(current, {})
            bucket[ticker] = bucket.get(ticker, 0.0) + float(value)
    return out


def load_holdings(wb, sheet: str = "Asset PM") -> dict[str, float]:
    """{ticker: market value} pooled across every account."""
    pooled: dict[str, float] = {}
    for holdings in load_holdings_by_account(wb, sheet).values():
        for ticker, value in holdings.items():
            pooled[ticker] = pooled.get(ticker, 0.0) + value
    return pooled


def is_deferred(account: str, taxable: set[str], deferred: set[str]) -> bool:
    if account in deferred:
        return True
    if account in taxable:
        return False
    return any(hint in account.upper() for hint in DEFERRED_HINTS)


def annualized_vol(series: dict, window: int = 13) -> float | None:
    """Realized volatility over the trailing `window` returns."""
    dates = sorted(series)[-(window + 1):]
    if len(dates) < 3:
        return None
    rets = [series[dates[i]] / series[dates[i - 1]] - 1 for i in range(1, len(dates))]
    vol = st.pstdev(rets) * math.sqrt(WEEKS_PER_YEAR)
    return vol if vol > 0 else None


def inverse_vol_weights(vols: dict[str, float]) -> dict[str, float]:
    """Weights proportional to 1/vol, normalized. Equal risk if uncorrelated."""
    inv = {t: 1.0 / v for t, v in vols.items()}
    total = sum(inv.values())
    return {t: x / total for t, x in inv.items()}


def risk_contributions(weights: dict[str, float],
                       vols: dict[str, float]) -> dict[str, float]:
    """Share of standalone risk per name. Ignores correlation — a first-order
    approximation, enough to expose a sizing mismatch, not a covariance
    model."""
    raw = {t: weights[t] * vols[t] for t in weights}
    total = sum(raw.values())
    return {t: v / total for t, v in raw.items()} if total else {t: 0.0 for t in weights}


def build_report(holdings: dict[str, float], prices: dict[str, dict],
                 window: int = 13, label: str = "", deferred: bool | None = None
                 ) -> dict[str, Any] | None:
    sized = {t: v for t, v in holdings.items() if t in prices}
    vols, skipped = {}, []
    for t in sized:
        v = annualized_vol(prices[t], window)
        vols[t] = v if v else skipped.append(t)
    for t in skipped:
        sized.pop(t, None)
        vols.pop(t, None)
    if len(sized) < 2:
        return None

    total = sum(sized.values())
    current = {t: v / total for t, v in sized.items()}
    target = inverse_vol_weights(vols)
    rc = risk_contributions(current, vols)

    rows = [
        dict(ticker=t, value=sized[t], weight=current[t], vol=vols[t],
             risk_share=rc[t], target_weight=target[t],
             trade=(target[t] - current[t]) * total)
        for t in sized
    ]
    rows.sort(key=lambda r: -r["risk_share"])
    return dict(
        label=label, deferred=deferred, total=total, window=window, rows=rows,
        blended_vol=sum(current[t] * vols[t] for t in sized),
        no_history=sorted(t for t in holdings if t not in prices),
        no_vol=sorted(skipped),
    )


def render(rep: dict[str, Any], show_missing: bool = True) -> str:
    head = rep["label"] or "SINGLE-NAME SLEEVE"
    tax = ""
    if rep.get("deferred") is True:
        tax = "  [tax-deferred — rebalancing realizes nothing]"
    elif rep.get("deferred") is False:
        tax = "  [taxable — trades realize gains]"
    out = [
        f"{head} — ${rep['total']:,.0f} across {len(rep['rows'])} names{tax}",
        f"blended volatility {rep['blended_vol']:.0%} · {rep['window']}-week window",
        "",
        f"{'Ticker':<8}{'value':>11}{'weight':>8}{'vol':>7}"
        f"{'risk':>8}{'target':>8}{'trade':>12}",
        "-" * 62,
    ]
    for r in rep["rows"]:
        out.append(
            f"{r['ticker']:<8}{r['value']:>11,.0f}{r['weight']:>8.1%}"
            f"{r['vol']:>7.0%}{r['risk_share']:>8.1%}"
            f"{r['target_weight']:>8.1%}{r['trade']:>+12,.0f}"
        )
    out.append("-" * 62)
    top = rep["rows"][0]
    out.append(f"Largest: {top['ticker']} — {top['weight']:.1%} of capital, "
               f"{top['risk_share']:.1%} of risk ({top['risk_share']/top['weight']:.1f}x).")
    if show_missing and rep["no_history"]:
        out.append(f"\nNo price history (not sized): {', '.join(rep['no_history'])}")
    return "\n".join(out)


FOOTER = """
Sizing only — no directional forecast. The measured edge is small and inside
the noise; weigh any taxable trade against a gain realized today."""


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--workbook", required=True, help="path to the portfolio .xlsx")
    ap.add_argument("--holdings-sheet", default="Asset PM")
    ap.add_argument("--vol-window", type=int, default=13,
                    help="trailing weeks for realized vol (default 13)")
    ap.add_argument("--pooled", action="store_true",
                    help="size across all accounts as one pool (NOT executable — "
                         "accounts cannot transfer to each other)")
    ap.add_argument("--taxable", default="", help="comma-separated account names to "
                                                  "force as taxable")
    ap.add_argument("--deferred", default="", help="comma-separated account names to "
                                                   "force as tax-deferred")
    ap.add_argument("--json", action="store_true", help="emit JSON instead of tables")
    args = ap.parse_args(argv)

    try:
        import openpyxl
    except ImportError:
        raise SystemExit("openpyxl required: uv sync")

    wb = openpyxl.load_workbook(args.workbook, data_only=True, read_only=True)
    prices = load_price_history(wb)
    taxable = {s.strip() for s in args.taxable.split(",") if s.strip()}
    forced = {s.strip() for s in args.deferred.split(",") if s.strip()}

    if args.pooled:
        reports = [build_report(load_holdings(wb, args.holdings_sheet), prices,
                                args.vol_window, "POOLED (not executable)")]
    else:
        reports = []
        for account, holdings in load_holdings_by_account(wb, args.holdings_sheet).items():
            rep = build_report(holdings, prices, args.vol_window, account,
                               is_deferred(account, taxable, forced))
            if rep:
                reports.append(rep)
        reports.sort(key=lambda r: -r["total"])

    reports = [r for r in reports if r]
    if not reports:
        raise SystemExit("no account has two or more names with price history")

    if args.json:
        print(json.dumps(reports, indent=2, default=str))
    else:
        print("\n\n".join(render(r, show_missing=args.pooled) for r in reports))
        print(FOOTER)
    return 0


if __name__ == "__main__":
    sys.exit(main())
