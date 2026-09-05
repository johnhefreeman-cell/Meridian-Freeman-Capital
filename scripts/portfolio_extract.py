"""Extract the whole book from the holdings workbook, reconciled account by account.

Everything the portfolio dashboard shows comes from here. The point of the
script is not convenience — it is that no figure on the dashboard is typed by
hand, so any of them can be re-derived from the workbook on demand.

**The reconciliation is the test.** Each account's holdings must sum to the
value the workbook's own ACCOUNT SUMMARY block reports for it, to the cent,
before anything is emitted. That catches the two failure modes that matter:
a row silently skipped (cash lines carry their label in column D and no ticker,
so a naive ticker filter drops them), and a margin loan quietly double-counted
(the joint account's holdings sum to gross; the summary reports net).

**Volatility is only available for single names.** The workbook carries weekly
price history on a per-name tab and none for funds, so risk shares are computed
within the single-name sleeve and say nothing about the rest of the book. A
single name with no tab — GOOGL, at the time of writing — is reported as
unmeasured rather than silently dropped from the denominator.

**Capital and risk shares use the same denominator** — the measured sleeve —
so the two are directly comparable. Sizing is per account, never pooled:
accounts cannot transfer to each other, so a target computed across the book
is not executable. See scripts/risk_sizing.py.

Usage:
    uv run python scripts/portfolio_extract.py --workbook path/to/Portfolio.xlsx
    uv run python scripts/portfolio_extract.py --workbook p.xlsx --json out.json
"""

from __future__ import annotations

import argparse
import json
import math
import statistics as st
import sys
from typing import Any

import openpyxl

WEEKS_PER_YEAR = 52.0
VOL_WINDOW = 13
HOLDINGS_SHEET = "Asset PM"
PRICE_PREFIX = "HMM "
PRICE_SKIP = {"HMM Data", "HMM Backtest"}
DATE_COL, PRICE_COL = 7, 8
SUMMARY_FIRST_ROW, SUMMARY_LAST_ROW = 5, 17

# Column offsets within an account block on the holdings sheet.
C_CLASS, C_SUBCLASS, C_LABEL, C_TICKER = 0, 1, 3, 4
C_PRICE, C_COST, C_HIGH, C_BELOW, C_SHARES, C_VALUE = 5, 6, 7, 8, 9, 10
C_ACCTPCT, C_GOAL, C_DIFF, C_VSCOST, C_GAIN = 11, 12, 13, 14, 15

# Bucket assignment. Single names are the only bucket with price history, and
# so the only one the risk arithmetic can speak about.
SINGLE_NAMES = {"MRVL", "CDNS", "AMZN", "GOOGL", "AMAT", "KLAC", "OUST", "SNDK",
                "AVGO", "BRK.B", "LLY", "META", "NVDA", "AAPL", "ORCL", "CRWD",
                "TSLA", "MSFT", "ASML"}
CASH_LIKE = {"CASH", "FDRXX", "SGOV", "SPAXX"}
CRYPTO = {"BTC", "ETH"}
REAL_ASSETS = {"GLD", "IBIT", "DBMF", "VNQ", "VNQI", "SPCX"}
BONDS = {"BND", "TIP", "VBTLX", "VWEAX", "MUB", "AGG"}
INTERNATIONAL = {"VXUS", "VTMGX", "VEMAX", "FNDE", "SCHY", "COPY", "VIGI",
                 "VNQI", "ASML"}

# Rows in the Crypto block carry a coin name in a merged cell rather than a
# ticker, so they are addressed by row index. Update if the sheet is restructured.
CRYPTO_ROWS = {168: "BTC", 169: "BTC", 170: "ETH", 171: "ETH"}

ACCOUNT_ALIASES = {"Tre Brokerage Account": "Tre Brokerage"}


def bucket(ticker: str) -> str:
    if ticker in CASH_LIKE:
        return "Cash"
    if ticker in CRYPTO:
        return "Crypto"
    if ticker in SINGLE_NAMES:
        return "Single names"
    if ticker in BONDS:
        return "Fixed income"
    if ticker in REAL_ASSETS:
        return "Real assets & alts"
    return "Diversified funds"


def _num(cell: Any) -> float | None:
    return float(cell) if isinstance(cell, (int, float)) else None


def account_summary(rows: list[tuple]) -> tuple[list[dict], dict]:
    """The workbook's own account totals — the figures everything reconciles to."""
    accounts = []
    for row in rows[SUMMARY_FIRST_ROW:SUMMARY_LAST_ROW]:
        name = row[1]
        if not isinstance(name, str) or name.strip().upper().startswith("TOTAL"):
            break
        accounts.append(dict(account=name.strip(), value=_num(row[5]),
                             share=_num(row[6]), gain=_num(row[7]),
                             divYield=_num(row[8])))
    total = dict(value=_num(rows[SUMMARY_LAST_ROW][5]),
                 gain=_num(rows[SUMMARY_LAST_ROW][7]),
                 divYield=_num(rows[SUMMARY_LAST_ROW][8]))
    return accounts, total


def load_holdings(rows: list[tuple], known: set[str]) -> list[dict]:
    """Every position row, tagged with the account whose block it sits in."""
    out: list[dict] = []
    current: str | None = None
    klass: str | None = None

    for i, row in enumerate(rows):
        if i < SUMMARY_LAST_ROW + 9:
            continue
        head = row[C_CLASS] if isinstance(row[C_CLASS], str) else None
        if head:
            name = ACCOUNT_ALIASES.get(head.strip(), head.strip())
            if name in known:
                current, klass = name, None
                continue
            if head.strip() != "Asset Classes":
                klass = head.strip()
        if isinstance(row[C_SUBCLASS], str) and row[C_SUBCLASS].strip():
            klass = row[C_SUBCLASS].strip()
        if current is None or len(row) <= C_VALUE:
            continue

        value = _num(row[C_VALUE])
        ticker = row[C_TICKER]

        if i in CRYPTO_ROWS and current == "Crypto":
            ticker = CRYPTO_ROWS[i]
        elif not isinstance(ticker, str) or not ticker.strip():
            # Cash lines name themselves in column D and carry no ticker.
            label = row[C_LABEL] if isinstance(row[C_LABEL], str) else None
            if label and value and (
                    "cash" in label.lower() or label.strip() in CASH_LIKE):
                out.append(dict(account=current, cls="Cash", ticker="CASH",
                                value=value, acctPct=_num(row[C_ACCTPCT]),
                                price=None, cost=None, high=None, belowHigh=None,
                                shares=None, vsCost=None, gain=None))
            continue

        ticker = ticker.strip()
        if ticker == "Ticker" or value is None:
            continue
        out.append(dict(account=current, cls=klass, ticker=ticker,
                        price=_num(row[C_PRICE]), cost=_num(row[C_COST]),
                        high=_num(row[C_HIGH]), belowHigh=_num(row[C_BELOW]),
                        shares=_num(row[C_SHARES]), value=value,
                        acctPct=_num(row[C_ACCTPCT]), goal=_num(row[C_GOAL]),
                        vsCost=_num(row[C_VSCOST]),
                        gain=_num(row[C_GAIN]) if len(row) > C_GAIN else None))
    return out


def realized_vol(wb, window: int = VOL_WINDOW) -> dict[str, float]:
    """Annualized volatility of the last `window` weekly returns, per name."""
    vols: dict[str, float] = {}
    for sheet in wb.sheetnames:
        if not sheet.startswith(PRICE_PREFIX) or sheet in PRICE_SKIP:
            continue
        series = {}
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if len(row) <= PRICE_COL:
                continue
            date, price = row[DATE_COL], row[PRICE_COL]
            if date is not None and isinstance(price, (int, float)) and price > 0:
                series[date] = float(price)
        if len(series) < window + 2:
            continue
        dates = sorted(series)[-(window + 1):]
        rets = [series[dates[i]] / series[dates[i - 1]] - 1
                for i in range(1, len(dates))]
        vols[sheet[len(PRICE_PREFIX):]] = st.pstdev(rets) * math.sqrt(WEEKS_PER_YEAR)
    return vols


def reconcile(accounts: list[dict], holdings: list[dict],
              margin: float, tol: float = 0.01) -> list[str]:
    """Every account's holdings must sum to its reported value.

    The joint account is the exception by construction: its holdings sum to
    gross and the summary reports net, so the margin loan closes the gap.
    """
    problems = []
    for acct in accounts:
        held = sum(h["value"] for h in holdings if h["account"] == acct["account"])
        expected = acct["value"]
        if abs(held - expected) > tol:
            if abs(held + margin - expected) <= tol:
                continue  # levered account: gross vs net
            problems.append(
                f"{acct['account']}: holdings {held:,.2f} vs reported "
                f"{expected:,.2f} (off by {held - expected:,.2f})")
    return problems


def build(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if HOLDINGS_SHEET not in wb.sheetnames:
        raise SystemExit(f"holdings sheet {HOLDINGS_SHEET!r} not in workbook")
    rows = list(wb[HOLDINGS_SHEET].iter_rows(values_only=True))

    accounts, total = account_summary(rows)
    margin = _num(rows[20][5]) or 0.0
    holdings = load_holdings(rows, {a["account"] for a in accounts})
    for h in holdings:
        h["bucket"] = bucket(h["ticker"])
        h["intl"] = h["ticker"] in INTERNATIONAL

    problems = reconcile(accounts, holdings, margin)
    if problems:
        raise SystemExit("reconciliation failed:\n  " + "\n  ".join(problems))

    vols = realized_vol(wb)
    gross = sum(h["value"] for h in holdings)

    per_account = []
    for acct in accounts:
        held = [h for h in holdings if h["account"] == acct["account"]]
        sleeve = [h for h in held if h["bucket"] == "Single names"]
        measured = [h for h in sleeve if vols.get(h["ticker"])]
        cap_den = sum(h["value"] for h in measured)
        risk_den = sum(h["value"] * vols[h["ticker"]] for h in measured)
        names = []
        for h in sorted(sleeve, key=lambda x: -x["value"]):
            vol = vols.get(h["ticker"])
            names.append(dict(
                ticker=h["ticker"], value=h["value"], vol=vol,
                capPct=(h["value"] / cap_den) if (vol and cap_den) else None,
                riskPct=(h["value"] * vol / risk_den) if (vol and risk_den) else None,
                belowHigh=h["belowHigh"], vsCost=h["vsCost"]))
        per_account.append(dict(
            account=acct["account"], value=acct["value"], share=acct["share"],
            gain=acct["gain"], divYield=acct["divYield"],
            gross=sum(h["value"] for h in held), nHold=len(held),
            singleName=sum(h["value"] for h in sleeve),
            measuredSleeve=cap_den,
            cash=sum(h["value"] for h in held if h["bucket"] == "Cash"),
            names=names))

    by_ticker: dict[str, dict] = {}
    for h in holdings:
        e = by_ticker.setdefault(h["ticker"], dict(
            ticker=h["ticker"], value=0.0, gain=0.0, accounts=[],
            bucket=h["bucket"], intl=h["intl"]))
        e["value"] += h["value"]
        e["gain"] += h["gain"] or 0.0
        e["accounts"].append(dict(account=h["account"], value=h["value"]))
        for k in ("price", "high", "belowHigh", "vsCost"):
            if h.get(k) is not None:
                e.setdefault(k, h[k])
    look = sorted(by_ticker.values(), key=lambda x: -x["value"])
    for e in look:
        e["accounts"].sort(key=lambda a: -a["value"])
        e["pct"] = e["value"] / gross
        e["vol"] = vols.get(e["ticker"])

    intl = sum(h["value"] for h in holdings if h["intl"])
    buckets: dict[str, float] = {}
    for h in holdings:
        buckets[h["bucket"]] = buckets.get(h["bucket"], 0.0) + h["value"]

    return dict(
        gross=gross, net=total["value"], loan=margin, gain=total["gain"],
        divYield=total["divYield"], marginRate=_num(rows[21][5]),
        intl=intl, intlPct=intl / gross,
        buckets=[dict(name=k, value=v, pct=v / gross)
                 for k, v in sorted(buckets.items(), key=lambda kv: -kv[1])],
        accounts=per_account, positions=look, vols=vols)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--json", help="write the full dataset here")
    args = ap.parse_args()

    d = build(args.workbook)
    if args.json:
        with open(args.json, "w") as fh:
            json.dump(d, fh, indent=1, default=str)

    print(f"Gross holdings   {d['gross']:>14,.2f}")
    print(f"Margin loan      {d['loan']:>14,.2f}")
    print(f"Net equity       {d['net']:>14,.2f}   "
          f"(reconciles across {len(d['accounts'])} accounts)")
    print(f"Unrealised       {d['gain']:>14,.2f}\n")

    for b in d["buckets"]:
        print(f"  {b['name']:<22}{b['value']:>13,.0f}{b['pct']:>8.1%}")

    sleeve = [p for p in d["positions"] if p["bucket"] == "Single names"]
    sleeve_v = sum(p["value"] for p in sleeve)
    unmeasured = [p for p in sleeve if not p["vol"]]
    print(f"\n{len(sleeve)} single names, {sleeve_v:,.0f} "
          f"({sleeve_v / d['gross']:.1%} of gross)")
    if unmeasured:
        print("  no price series, absent from every risk figure: " +
              ", ".join(f"{p['ticker']} {p['value']:,.0f}" for p in unmeasured))

    for a in d["accounts"]:
        if not any(n["riskPct"] for n in a["names"]):
            continue
        print(f"\n{a['account']} — sleeve {a['singleName']:,.0f}, "
              f"measured {a['measuredSleeve']:,.0f}")
        print(f"  {'name':<8}{'capital':>9}{'risk':>8}{'ratio':>8}{'vol':>7}")
        for n in a["names"]:
            if n["riskPct"] is None:
                print(f"  {n['ticker']:<8}{'—':>9}{'—':>8}{'—':>8}{'—':>7}"
                      f"   {n['value']:,.0f} unmeasured")
                continue
            print(f"  {n['ticker']:<8}{n['capPct']:>9.1%}{n['riskPct']:>8.1%}"
                  f"{n['riskPct'] / n['capPct']:>7.2f}x{n['vol']:>7.0%}")


if __name__ == "__main__":
    main()
